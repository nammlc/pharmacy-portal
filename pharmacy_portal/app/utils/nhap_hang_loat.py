"""
Nhập liệu hàng loạt (bulk import) cho Danh mục thuốc / Nhà thuốc BV
từ file CSV hoặc Excel (.xlsx), để nhập nhanh nhiều thuốc cùng lúc
thay vì phải thêm từng thuốc một qua form.

Cách dùng (xem route app/routes/admin_nhap_hang_loat.py):
    rows = doc_file_bang(file_storage)
    ket_qua = nhap_danh_muc_thuoc(rows)   # hoặc nhap_nha_thuoc_bv(rows)
"""

from __future__ import annotations
import csv
import io
import re
import unicodedata

from openpyxl import load_workbook

from app import db
from app.models.models import NhomThuoc, HoatChat, DanhMucThuoc, NhaThuocBV
from app.utils.lam_sach_html import lam_sach_html


# ---------------------------------------------------------------------------
# Đọc file (CSV hoặc Excel) thành list[dict] - key là tên cột đã chuẩn hoá
# ---------------------------------------------------------------------------

def _chuan_hoa_ten_cot(ten) -> str:
    """'Tên biệt dược' / 'TÊN-BIỆT DƯỢC' / 'ten_biet_duoc' -> 'ten_biet_duoc'."""
    ten = str(ten or "").strip().lower()
    ten = unicodedata.normalize("NFD", ten)
    ten = "".join(ch for ch in ten if unicodedata.category(ch) != "Mn")  # bỏ dấu
    ten = ten.replace("đ", "d")
    ten = re.sub(r"[^a-z0-9]+", "_", ten).strip("_")
    return ten


def doc_file_bang(file_storage) -> list[dict]:
    """
    Đọc file .csv hoặc .xlsx thành list[dict], dòng đầu tiên là tên cột.
    Bỏ qua các dòng trống hoàn toàn.
    """
    ten_file = (file_storage.filename or "").lower()

    if ten_file.endswith(".xlsx"):
        wb = load_workbook(file_storage.stream, data_only=True, read_only=True)
        sheet = wb.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header = [_chuan_hoa_ten_cot(c) for c in next(rows_iter)]
        except StopIteration:
            return []
        ket_qua = []
        for row in rows_iter:
            if row is None or all(o is None or str(o).strip() == "" for o in row):
                continue
            item = {header[i]: (row[i] if i < len(row) else None) for i in range(len(header))}
            ket_qua.append(item)
        return ket_qua

    # CSV - thử vài encoding phổ biến (utf-8-sig, cp1258 kiểu Excel Việt hoá cũ)
    raw = file_storage.stream.read()
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1258", "cp1252"):
        try:
            text = raw.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")

    # Tự dò dấu phân cách (, hoặc ;) - file Excel Việt Nam hay xuất ra dấu ;
    mau = text[:2048]
    dau_phay_cach = ";" if mau.count(";") > mau.count(",") else ","

    reader = csv.reader(io.StringIO(text), delimiter=dau_phay_cach)
    rows = list(reader)
    if not rows:
        return []
    header = [_chuan_hoa_ten_cot(c) for c in rows[0]]
    ket_qua = []
    for row in rows[1:]:
        if not row or all(not str(o).strip() for o in row):
            continue
        item = {header[i]: (row[i] if i < len(row) else None) for i in range(len(header))}
        ket_qua.append(item)
    return ket_qua


# ---------------------------------------------------------------------------
# Lấy hoặc tạo mới nhóm thuốc / hoạt chất (không tạo trùng, so sánh không phân biệt hoa-thường)
# ---------------------------------------------------------------------------

def _lay_hoac_tao_nhom_thuoc(ten: str, loai: str) -> NhomThuoc | None:
    ten = (ten or "").strip()
    if not ten:
        return None
    nhom = NhomThuoc.query.filter(
        db.func.lower(NhomThuoc.ten_nhom) == ten.lower(), NhomThuoc.loai == loai
    ).first()
    if nhom:
        return nhom
    nhom = NhomThuoc(ten_nhom=ten, loai=loai)
    db.session.add(nhom)
    db.session.flush()
    return nhom


def _lay_hoac_tao_hoat_chat_list(chuoi_hoat_chat: str) -> list[HoatChat]:
    """'Paracetamol + Cafein' hoặc 'Paracetamol, Cafein' -> [HoatChat, HoatChat]."""
    chuoi_hoat_chat = (chuoi_hoat_chat or "").strip()
    if not chuoi_hoat_chat:
        return []
    ten_list = [t.strip() for t in re.split(r"[+,;/]", chuoi_hoat_chat) if t.strip()]
    ket_qua = []
    for ten in ten_list:
        hc = HoatChat.query.filter(db.func.lower(HoatChat.ten_hoat_chat) == ten.lower()).first()
        if not hc:
            hc = HoatChat(ten_hoat_chat=ten)
            db.session.add(hc)
            db.session.flush()
        ket_qua.append(hc)
    return ket_qua


# ---------------------------------------------------------------------------
# Kết quả trả về cho route hiển thị báo cáo
# ---------------------------------------------------------------------------

class KetQuaNhap:
    def __init__(self):
        self.so_thanh_cong = 0
        self.so_loi = 0
        self.chi_tiet_loi: list[str] = []  # ["Dòng 3: thiếu tên biệt dược", ...]

    def them_loi(self, dong: int, ly_do: str):
        self.so_loi += 1
        self.chi_tiet_loi.append(f"Dòng {dong}: {ly_do}")


# ---------------------------------------------------------------------------
# Nhập Danh mục thuốc
# ---------------------------------------------------------------------------

def nhap_danh_muc_thuoc(rows: list[dict]) -> KetQuaNhap:
    kq = KetQuaNhap()
    for idx, row in enumerate(rows, start=2):  # dòng 1 là header
        ten_biet_duoc = str(row.get("ten_biet_duoc") or "").strip()
        if not ten_biet_duoc:
            kq.them_loi(idx, "thiếu 'ten_biet_duoc' (tên biệt dược) - bỏ qua dòng này")
            continue
        try:
            thuoc = DanhMucThuoc.query.filter(
                db.func.lower(DanhMucThuoc.ten_biet_duoc) == ten_biet_duoc.lower()
            ).first()
            if not thuoc:
                thuoc = DanhMucThuoc(ten_biet_duoc=ten_biet_duoc)
                db.session.add(thuoc)

            nhom = _lay_hoac_tao_nhom_thuoc(row.get("nhom_thuoc"), "danh_muc_thuoc")
            if nhom:
                thuoc.nhom_thuoc_id = nhom.id

            hoat_chat_list = _lay_hoac_tao_hoat_chat_list(row.get("hoat_chat"))
            if hoat_chat_list:
                thuoc.hoat_chat_list = hoat_chat_list

            if row.get("thanh_phan"):
                thuoc.thanh_phan = lam_sach_html(str(row["thanh_phan"]))
            if row.get("chi_dinh"):
                thuoc.chi_dinh = lam_sach_html(str(row["chi_dinh"]))
            if row.get("chong_chi_dinh"):
                thuoc.chong_chi_dinh = lam_sach_html(str(row["chong_chi_dinh"]))
            if row.get("cach_dung_lieu_dung"):
                thuoc.cach_dung_lieu_dung = lam_sach_html(str(row["cach_dung_lieu_dung"]))
            if row.get("link_chi_tiet"):
                thuoc.link_chi_tiet = str(row["link_chi_tiet"]).strip()

            db.session.flush()
            kq.so_thanh_cong += 1
        except Exception as e:
            db.session.rollback()
            kq.them_loi(idx, f"lỗi khi lưu '{ten_biet_duoc}' - {e}")

    db.session.commit()
    return kq


# ---------------------------------------------------------------------------
# Nhập Nhà thuốc BV
# ---------------------------------------------------------------------------

def nhap_nha_thuoc_bv(rows: list[dict]) -> KetQuaNhap:
    kq = KetQuaNhap()
    for idx, row in enumerate(rows, start=2):
        ten_biet_duoc = str(row.get("ten_biet_duoc") or "").strip()
        if not ten_biet_duoc:
            kq.them_loi(idx, "thiếu 'ten_biet_duoc' (tên biệt dược) - bỏ qua dòng này")
            continue
        try:
            thuoc = NhaThuocBV.query.filter(
                db.func.lower(NhaThuocBV.ten_biet_duoc) == ten_biet_duoc.lower()
            ).first()
            if not thuoc:
                thuoc = NhaThuocBV(ten_biet_duoc=ten_biet_duoc)
                db.session.add(thuoc)

            nhom = _lay_hoac_tao_nhom_thuoc(row.get("nhom_thuoc"), "nha_thuoc_bv")
            if nhom:
                thuoc.nhom_thuoc_id = nhom.id

            hoat_chat_list = _lay_hoac_tao_hoat_chat_list(row.get("hoat_chat"))
            if hoat_chat_list:
                thuoc.hoat_chat_list = hoat_chat_list

            if row.get("link_tham_khao"):
                thuoc.link_tham_khao = str(row["link_tham_khao"]).strip()

            db.session.flush()
            kq.so_thanh_cong += 1
        except Exception as e:
            db.session.rollback()
            kq.them_loi(idx, f"lỗi khi lưu '{ten_biet_duoc}' - {e}")

    db.session.commit()
    return kq
